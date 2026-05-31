"""
Data Loader - Parser file Excel PDRB, Penduduk, Kode Wilayah
Menghasilkan struktur data terpadu untuk semua analisis.
"""

import os
import pickle
import warnings
import numpy as np
import pandas as pd
import openpyxl
import streamlit as st

warnings.filterwarnings("ignore")

# Import config (path-safe)
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config


# ──────────────────────────────────────────────────────────────────────────────
# KODE WILAYAH
# ──────────────────────────────────────────────────────────────────────────────

def load_kode_wilayah(filepath=None):
    """
    Memuat kode wilayah, nama, dan kelompok pembangunan.
    Returns: dict { kode_str: { 'name': ..., 'kelompok': ... } }
    """
    fp = filepath or config.KODE_WILAYAH_FILE
    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb.active
    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        kode, nama, kelompok = row[0], row[1], row[2]
        if kode:
            result[str(int(kode))] = {
                "name": nama or "",
                "kelompok": kelompok or "",
            }
    wb.close()
    return result


# ──────────────────────────────────────────────────────────────────────────────
# PENDUDUK
# ──────────────────────────────────────────────────────────────────────────────

def load_penduduk(filepath=None):
    """
    Memuat data jumlah penduduk triwulanan.
    Returns: dict { kode_str: { 'YYYYQN': jiwa_ribu, ... } }
    """
    fp = filepath or config.PENDUDUK_FILE
    wb = openpyxl.load_workbook(fp, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Baris 1: header tahun (dengan None untuk kolom sub-quarter)
    # Baris 2: header quarter (Q1/Q2/Q3/Q4)
    year_row = rows[0]
    q_row = rows[1]

    # Bangun mapping kolom → "YYYYQn"
    col_periods = {}
    cur_year = None
    for c_idx, (yr, qt) in enumerate(zip(year_row, q_row)):
        if isinstance(yr, (int, float)) and yr > 2000:
            cur_year = int(yr)
        if qt in ("Q1", "Q2", "Q3", "Q4") and cur_year:
            qnum = int(qt[1])
            col_periods[c_idx] = f"{cur_year}Q{qnum}"

    result = {}
    for row in rows[2:]:
        kode = row[0]
        if not isinstance(kode, (int, float)):
            continue
        kode_str = str(int(kode))
        vals = {}
        for c_idx, period in col_periods.items():
            v = row[c_idx]
            if isinstance(v, (int, float)):
                vals[period] = float(v)
        result[kode_str] = vals

    wb.close()
    return result


# ──────────────────────────────────────────────────────────────────────────────
# PDRB – PARSER UTAMA
# ──────────────────────────────────────────────────────────────────────────────

def _parse_col_headers(rows_4_5):
    """
    Dari baris header (row 4 & row 5 dari sheet), bangun daftar ordered periods.
    Setiap periode = { 'col_idx': n, 'year': 2020, 'quarter': 1, 'period': '2020Q1',
                       'is_annual': False }
    """
    year_row, q_row = rows_4_5
    periods = []
    cur_year = None
    for c_idx, (yr, qt) in enumerate(zip(year_row, q_row)):
        if isinstance(yr, str) and yr.startswith("Triwulanan"):
            try:
                cur_year = int(yr.split()[-1])
            except Exception:
                pass
        elif isinstance(yr, (int, float)) and 2005 <= yr <= 2035:
            if qt is None:
                # kolom tahunan – skip
                pass
            cur_year = int(yr)
        if qt in ("I", "II", "III", "IV") and cur_year:
            qmap = {"I": 1, "II": 2, "III": 3, "IV": 4}
            qnum = qmap[qt]
            periods.append({
                "col_idx": c_idx,
                "year": cur_year,
                "quarter": qnum,
                "period": f"{cur_year}Q{qnum}",
                "is_annual": False,
            })
    # deduplicate periods (same period string, keep first)
    seen = {}
    result = []
    for p in periods:
        if p["period"] not in seen:
            seen[p["period"]] = True
            result.append(p)
    return result


def _build_sector_map(data_rows):
    """
    Dari baris data, bangun mapping: kode_kat → { name, parent, row_idx }
    """
    sectors = {}
    for r_idx, row in enumerate(data_rows):
        kat_a  = row[0]  # Huruf kategori (A, B, C, ...)
        kat_b  = row[1]  # Nomor sub (1, 2, ...) or teks nama
        name   = row[2]  # Nama panjang (None jika kategori utama)
        kode_s = row[3]  # Kode string: A, A1, A01, B, ...

        if not kode_s or str(kode_s).strip() in ("", "Sub", "Kat", "-1", "-2"):
            continue
        kode_s = str(kode_s).strip()

        # Tentukan nama dan level
        if name:
            label = str(name).strip()
        elif kat_b and isinstance(kat_b, str) and not kat_b.isdigit():
            label = str(kat_b).strip()
        elif kat_a and isinstance(kat_a, str):
            label = str(kat_a).strip() if not name else str(name).strip()
        else:
            label = kode_s

        # Tentukan parent
        parent = None
        if len(kode_s) > 1:
            # Sub-sektor → cari parent dengan kode 1 huruf
            parent_candidates = [s for s in sectors if len(s) == 1 and kode_s.startswith(s)]
            if parent_candidates:
                parent = parent_candidates[-1]

        sectors[kode_s] = {
            "name": label,
            "parent": parent,
            "row_idx": r_idx,
        }
    return sectors


def _extract_values(data_rows, row_idx, col_periods):
    """Ekstrak nilai numerik dari satu baris."""
    row = data_rows[row_idx]
    vals = {}
    for p in col_periods:
        c = p["col_idx"]
        if c < len(row):
            v = row[c]
            if isinstance(v, (int, float)) and not (v != v):  # not NaN
                vals[p["period"]] = float(v)
            else:
                vals[p["period"]] = None
        else:
            vals[p["period"]] = None
    return vals


def _parse_sheet(ws, sheet_name):
    """
    Parse satu sheet PDRB → dict dengan kunci 'adhb' dan 'adhk'.
    Struktur masing-masing:
      {
        'periods': [ '2010Q1', ... ],
        'sectors': { kode: { 'name', 'parent', 'values': {period: val} } },
        'total':   { period: val },
        'total_nonmigas': { period: val },
      }
    """
    all_rows = list(ws.iter_rows(values_only=True))

    # Temukan baris TABEL 1 dan TABEL 2
    tabel1_row = None
    tabel2_row = None
    for i, row in enumerate(all_rows):
        if row[0] and "TABEL 1" in str(row[0]):
            tabel1_row = i
        if row[0] and "TABEL 2" in str(row[0]):
            tabel2_row = i

    if tabel1_row is None:
        return None

    result = {}

    for tbl_key, tbl_start in [("adhb", tabel1_row), ("adhk", tabel2_row)]:
        if tbl_start is None:
            continue

        # Header di tbl_start+3 dan tbl_start+4 (row 4 & 5 relatif ke tabel)
        hdr_idx = tbl_start + 3
        q_idx   = tbl_start + 4

        if hdr_idx >= len(all_rows):
            continue

        col_periods = _parse_col_headers([all_rows[hdr_idx], all_rows[q_idx]])

        if not col_periods:
            continue

        # Data mulai dari tbl_start+8 (row 9 = index 8 relatif)
        data_start = tbl_start + 8

        # Cari batas akhir tabel
        if tbl_key == "adhb" and tabel2_row:
            data_end = tabel2_row
        else:
            data_end = len(all_rows)

        data_rows = all_rows[data_start:data_end]

        # Bangun sector map
        sector_map = _build_sector_map(data_rows)

        # Cari baris PDRB total dan total non-migas
        total_vals = {}
        total_nm_vals = {}
        for r_idx, row in enumerate(data_rows):
            if row[1] and "PRODUK DOMESTIK REGIONAL BRUTO" in str(row[1]).upper():
                total_vals = _extract_values(data_rows, r_idx, col_periods)
            if row[0] and "PRODUK DOMESTIK REGIONAL BRUTO TANPA MIGAS" in str(row[0]).upper():
                total_nm_vals = _extract_values(data_rows, r_idx, col_periods)

        # Bangun nilai per sektor
        sectors_out = {}
        for kode, meta in sector_map.items():
            vals = _extract_values(data_rows, meta["row_idx"], col_periods)
            sectors_out[kode] = {
                "name": meta["name"],
                "parent": meta["parent"],
                "values": vals,
            }

        # Filter periode: hanya yang punya total > 0 (ada data nyata)
        valid_periods = [
            p["period"] for p in col_periods
            if total_vals.get(p["period"]) is not None
            and float(total_vals.get(p["period"]) or 0) > 0
        ]
        # Jika total tidak ditemukan, fallback ke periode dengan setidaknya 1 sektor punya nilai
        if not valid_periods and total_vals:
            valid_periods = [
                p["period"] for p in col_periods
                if any(
                    sectors_out.get(ks, {}).get("values", {}).get(p["period"])
                    for ks in sectors_out
                )
            ]

        result[tbl_key] = {
            "periods": valid_periods,
            "sectors": sectors_out,
            "total": {p: v for p, v in total_vals.items()
                      if v is not None and float(v or 0) > 0},
            "total_nonmigas": total_nm_vals,
        }

    return result


@st.cache_data(show_spinner="Memuat data PDRB... (proses satu kali)")
def load_all_pdrb(filepath=None, force_reload=False):
    """
    Load & cache semua data PDRB dari file Excel.
    Returns: dict { kode_str: { 'adhb': {...}, 'adhk': {...} } }
    """
    fp = filepath or config.PDRB_FILE

    # Cek cache pickle
    cache_fp = config.CACHE_FILE
    if not force_reload and os.path.exists(cache_fp):
        mod_time_excel = os.path.getmtime(fp)
        mod_time_cache = os.path.getmtime(cache_fp)
        if mod_time_cache > mod_time_excel:
            with open(cache_fp, "rb") as f:
                return pickle.load(f)

    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    pdrb_data = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.isdigit():
            continue
        ws = wb[sheet_name]
        parsed = _parse_sheet(ws, sheet_name)
        if parsed:
            pdrb_data[sheet_name] = parsed

    wb.close()

    # Simpan cache
    os.makedirs(os.path.dirname(cache_fp), exist_ok=True)
    with open(cache_fp, "wb") as f:
        pickle.dump(pdrb_data, f)

    return pdrb_data


# ──────────────────────────────────────────────────────────────────────────────
# HELPER: BUAT DATAFRAME DARI DATA PDRB
# ──────────────────────────────────────────────────────────────────────────────

def get_period_list(pdrb_data, kode="3300", tabel="adhb"):
    """Ambil daftar periode yang tersedia untuk wilayah tertentu."""
    try:
        return pdrb_data[kode][tabel]["periods"]
    except Exception:
        return []


def get_sectors_df(pdrb_data, kode, tabel="adhb"):
    """
    Buat DataFrame: index=sektor_kode, columns=period, values=nilai.
    Juga tambahkan kolom metadata: name, parent, level.
    """
    if kode not in pdrb_data or tabel not in pdrb_data[kode]:
        return pd.DataFrame()

    tbl = pdrb_data[kode][tabel]
    periods = tbl["periods"]
    sectors = tbl["sectors"]

    records = []
    for kode_s, meta in sectors.items():
        row = {"kode": kode_s, "name": meta["name"], "parent": meta["parent"]}
        row["level"] = "main" if meta["parent"] is None else "sub"
        for p in periods:
            row[p] = meta["values"].get(p)
        records.append(row)

    df = pd.DataFrame(records)
    if not df.empty:
        df = df.set_index("kode")
    return df


def get_total_series(pdrb_data, kode, tabel="adhb"):
    """Ambil series PDRB total sebagai dict { period: nilai }."""
    try:
        return pdrb_data[kode][tabel]["total"]
    except Exception:
        return {}


def compute_growth(series_dict, mode="qtq"):
    """
    Hitung pertumbuhan dari dict { period: nilai }.
    mode: 'qtq' | 'yoy' | 'ctc'
    Returns: dict { period: pct_growth }
    """
    periods = sorted(series_dict.keys())
    vals = pd.Series({p: series_dict[p] for p in periods if series_dict[p] is not None})
    vals = vals.sort_index()

    if mode == "qtq":
        growth = vals.pct_change() * 100
    elif mode == "yoy":
        growth = vals.pct_change(periods=4) * 100
    elif mode == "ctc":
        # Kumulatif year-to-date vs year-to-date tahun lalu
        def get_year_q(p):
            y, q = int(p[:4]), int(p[5])
            return y, q
        cumulative = {}
        for p in vals.index:
            y, q = get_year_q(p)
            # Hitung kumulatif s.d. quarter ini dalam tahun y
            cum = sum(
                vals.get(f"{y}Q{qi}", 0) or 0
                for qi in range(1, q + 1)
                if f"{y}Q{qi}" in vals.index
            )
            cum_prev = sum(
                vals.get(f"{y-1}Q{qi}", 0) or 0
                for qi in range(1, q + 1)
                if f"{y-1}Q{qi}" in vals.index
            )
            if cum_prev and cum_prev != 0:
                cumulative[p] = (cum / cum_prev - 1) * 100
            else:
                cumulative[p] = None
        return cumulative
    else:
        growth = vals * 0  # zero

    return growth.to_dict()


def compute_distribution(sector_df, periods):
    """
    Hitung distribusi (%) setiap sektor utama terhadap total.
    Returns: DataFrame dengan nilai distribusi (%).
    """
    main_sectors = sector_df[sector_df["level"] == "main"] if "level" in sector_df.columns else sector_df
    period_cols = [p for p in periods if p in main_sectors.columns]
    data = main_sectors[period_cols].astype(float)
    total = data.sum(axis=0)
    dist = data.div(total, axis=1) * 100
    return dist


# ──────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────────
# CONVENIENCE LOADER
# ──────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner="Memuat semua data...")
def load_all_data():
    """
    Load semua data sekaligus: PDRB, Penduduk, Kode Wilayah.
    Returns: (pdrb_data, penduduk_data, kode_wilayah)
    """
    pdrb_data    = load_all_pdrb()
    penduduk_data = load_penduduk()
    kode_wilayah  = load_kode_wilayah()
    return pdrb_data, penduduk_data, kode_wilayah
